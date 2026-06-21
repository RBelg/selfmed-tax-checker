#!/usr/bin/env python3
"""
厚労省「セルフメディケーション税制対象品目一覧」xlsx を取得し、
照合・表示に使える data/medicines.json を生成する。

使い方:
    python scripts/build_medicines.py

月次更新: 厚労省ページ
  https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000124853.html
で最新版の xlsx URL を確認し、下記 SOURCES を差し替えてから再実行する。

データ送信は一切せず、生成物 medicines.json のみブラウザ側で読み込まれる。
"""
from __future__ import annotations

import io
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path

import openpyxl
import requests

# --- 取得元（月次で URL を差し替える） -----------------------------------
# 厚労省ページから「対象品目一覧」xlsx の直リンクを確認して更新すること。
# 非スイッチOTC は最新版が PDF のみの月がある。xlsx が出ている月のみここに追加する。
SOURCES = [
    {
        "type": "switch",
        "label": "スイッチOTC 令和8年6月公表",
        "url": "https://www.mhlw.go.jp/content/10800000/001705872.xlsx",
    },
    # 非スイッチOTC: 最新版が PDF のみのため当面未対応。
    # xlsx が公開されている月は下記の形で追加する:
    # {"type": "non-switch", "label": "非スイッチOTC 令和8年X月", "url": "https://www.mhlw.go.jp/content/..."},
]

HEADERS = {"User-Agent": "Mozilla/5.0 (selfmed-tax-checker data builder)"}

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "data" / "medicines.json"

# 列ヘッダ名 -> 出力キー（ヘッダ文字列の正規化後にマッチ）
COLUMN_ALIASES = {
    "販売名": "name",
    "有効成分名": "ingredient",
    "製造販売業者名": "maker",
    "ＪＡＮコード": "jan",
    "JANコード": "jan",
}

# 販売名から除去する括弧書き（リスク分類・注記など）
_BRACKET_RE = re.compile(r"[【〔\[(（].*?[】〕\])）]")
# 正規化で除去する記号・空白類
_STRIP_RE = re.compile(r"[\s　・,，.。/／\\\-－—–~〜=＝!！?？\"'’＇`*＊#＃&＆+＋:：;；]")


def normalize_name(name: str) -> str:
    """販売名 / Amazon商品名 を照合用キーに正規化する。
    ブラウザ側(JS)の normalize と同一ルールを維持すること。
    """
    if not name:
        return ""
    s = unicodedata.normalize("NFKC", str(name))
    s = _BRACKET_RE.sub("", s)
    s = s.lower()
    s = _STRIP_RE.sub("", s)
    return s


def normalize_header(value) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip().replace(" ", "")


def download(url: str) -> bytes:
    print(f"  download: {url}")
    r = requests.get(url, headers=HEADERS, timeout=120)
    r.raise_for_status()
    print(f"    {len(r.content):,} bytes")
    return r.content


def find_header_row(ws, max_scan: int = 20):
    """「販売名」を含む行をヘッダ行として特定し、(行番号, {出力キー: 列index0始まり}) を返す。"""
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        norm = [normalize_header(c) for c in row]
        if "販売名" in norm:
            mapping: dict[str, int] = {}
            for col_idx, h in enumerate(norm):
                key = COLUMN_ALIASES.get(h)
                if key and key not in mapping:
                    mapping[key] = col_idx
            return r_idx + 1, mapping  # iter_rows は0始まり、Excel行番号へ
    raise ValueError("ヘッダ行（『販売名』列）が見つかりませんでした")


def parse_xlsx(content: bytes, src_type: str) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, cols = find_header_row(ws)
    if "name" not in cols:
        wb.close()
        raise ValueError(f"販売名列を特定できません: cols={cols}")

    items = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = row[cols["name"]] if cols["name"] < len(row) else None
        if not name or not str(name).strip():
            continue
        name = str(name).strip()

        def get(key):
            idx = cols.get(key)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        items.append({
            "name": name,
            "name_norm": normalize_name(name),
            "ingredient": get("ingredient"),
            "maker": get("maker"),
            "type": src_type,
        })
    wb.close()
    return items


def dedupe(items: list[dict]) -> list[dict]:
    """name_norm で重複排除（同一販売名の容量違い等を1件に集約）。"""
    seen: dict[str, dict] = {}
    for it in items:
        k = it["name_norm"]
        if not k:
            continue
        if k not in seen:
            seen[k] = it
    return list(seen.values())


def main() -> int:
    all_items: list[dict] = []
    labels = []
    for src in SOURCES:
        print(f"[{src['type']}] {src['label']}")
        try:
            content = download(src["url"])
            items = parse_xlsx(content, src["type"])
            print(f"    parsed: {len(items)} 行")
            all_items.extend(items)
            labels.append(src["label"])
        except Exception as e:  # noqa: BLE001
            print(f"    ERROR: {e}", file=sys.stderr)
            return 1

    deduped = dedupe(all_items)
    print(f"合計 {len(all_items)} 行 -> 重複排除後 {len(deduped)} 件")

    out = {
        "updated": date.today().isoformat(),
        "source": " / ".join(labels),
        "count": len(deduped),
        "items": sorted(deduped, key=lambda x: x["name_norm"]),
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f"出力: {OUT_PATH} ({size_kb:.1f} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
